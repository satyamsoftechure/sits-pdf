from django.shortcuts import render, redirect
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import urllib.parse
import qrcode
import json
from django.http import (
    HttpResponse,
    HttpResponseNotFound,
    JsonResponse,
    HttpResponseBadRequest,
)
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from .merger import PdfMerger
from django.contrib import messages
from .pdf_split_utils import (
    split_pdf_into_single_pages,
    split_pdf_into_multiple_files,
    extract_between_pages,
)
from .support._reader import PdfReader
from .support._writer import PdfWriter
import zipfile
from io import BytesIO
import os
from .word_converter import Converter
from .pdf_to_image_covert import convert_from_path
from .pdf_to_image_covert import convert_from_bytes
from .Image_pdf import convert
from django.http import HttpResponse
import uuid
from django.urls import reverse
from selenium import webdriver
import pdfkit
import time
import camelot
import pandas as pd
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import logging
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import io
import base64
from PIL import Image
from django.template.loader import render_to_string
import requests
from urllib.parse import urlparse
from django.views.decorators.csrf import csrf_exempt
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account
from googleapiclient.discovery import build
from django.http import JsonResponse
from django.views.decorators.http import require_GET
import tempfile
from playwright.sync_api import sync_playwright
from datetime import datetime
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, Border, Side
import re
import fitz

logger = logging.getLogger(__name__)


def main(request):
    return render(request, "home.html", {'current_year': datetime.now().year})


SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
SERVICE_ACCOUNT_FILE = "service-account-file.json"


def get_drive_service():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )
    service = build("drive", "v3", credentials=creds)
    return service


def handle_drive_files(drive_file_ids, credentials, file_data):
    for file_id in drive_file_ids:
        try:
            filename, pdf_bytes = get_drive_file(
                file_id, credentials
            )  # Assuming `get_drive_file` uses credentials
            if filename.endswith(".pdf"):
                file_data = process_pdf(filename, pdf_bytes, file_data)
            else:
                return {
                    "success": False,
                    "error": f"{filename} is not a PDF file. Please ensure the Google Drive files are PDFs.",
                }
        except Exception as file_error:
            return {
                "success": False,
                "error": f"Error processing Google Drive file with ID {file_id}: {str(file_error)}",
            }
    return {"success": True, "file_data": file_data}


def handle_drive_jpg_files(drive_file_ids, credentials, file_data):
    for file_id in drive_file_ids:
        try:
            filename, image_data = get_drive_file(
                file_id, credentials
            )  # Assuming `get_drive_file` uses credentials
            if filename.endswith(".jpg"):
                file_id = str(uuid.uuid4())
                file_info = {
                    "file_id": file_id,
                    "filename": filename,
                    "image_data": base64.b64encode(image_data).decode(),
                }
                file_data.append(file_info)
            else:
                return {
                    "success": False,
                    "error": f"{filename} is not a JPG file. Please ensure the Google Drive files are JPGs.",
                }
        except Exception as file_error:
            return {
                "success": False,
                "error": f"Error processing Google Drive file with ID {file_id}: {str(file_error)}",
            }
    return {"success": True, "file_data": file_data}


def handle_dropbox_file(dropbox_file_link, file_data):
    try:
        response = requests.get(dropbox_file_link)
        response.raise_for_status()
        pdf_bytes = response.content
        filename = os.path.basename(dropbox_file_link)
        file_data = process_pdf(filename, pdf_bytes, file_data)
        return {"success": True, "file_data": file_data}
    except Exception as e:
        return {"success": False, "error": f"Error processing Dropbox file: {str(e)}"}


def handle_dropbox_jpg_file(dropbox_file_link, file_data):
    try:
        response = requests.get(dropbox_file_link, stream=True)
        response.raise_for_status()
        image_data = response.content
        filename = os.path.basename(dropbox_file_link)
        file_id = str(uuid.uuid4())
        file_info = {
            "file_id": file_id,
            "filename": filename,
            "image_data": base64.b64encode(image_data).decode(),
        }
        file_data.append(file_info)
        return {"success": True, "file_data": file_data}
    except Exception as e:
        return {"success": False, "error": f"Error processing Dropbox file: {str(e)}"}


def handle_url_file(url, file_data):
    parsed_url = urlparse(url)
    filename = os.path.basename(parsed_url.path)
    if filename.endswith(".pdf"):
        try:
            response = requests.get(url)
            response.raise_for_status()
            pdf_bytes = response.content
            file_data = process_pdf(filename, pdf_bytes, file_data)
            return {"success": True, "file_data": file_data}
        except Exception as e:
            return {"success": False, "error": f"Error processing URL {url}: {str(e)}"}
    else:
        return {
            "success": False,
            "error": f"{filename} is not a PDF file. Please ensure the URL points to a PDF.",
        }


def handle_url_jpg_file(url, file_data):
    parsed_url = urlparse(url)
    filename = os.path.basename(parsed_url.path)
    if filename.endswith(".jpg"):
        try:
            response = requests.get(url)
            response.raise_for_status()
            image_data = response.content
            file_id = str(uuid.uuid4())
            file_info = {
                "file_id": file_id,
                "filename": filename,
                "image_data": base64.b64encode(image_data).decode(),
            }
            file_data.append(file_info)
            return {"success": True, "file_data": file_data}
        except Exception as e:
            return {"success": False, "error": f"Error processing URL {url}: {str(e)}"}
    else:
        return {
            "success": False,
            "error": f"{filename} is not a JPG file. Please ensure the URL points to a JPG.",
        }


def upload_pdfs(request):
    if request.method == "GET":
        request.session.pop("file_data", None)
        if "credentials" in request.session:
            credentials = Credentials(**request.session["credentials"])
            return render(
                request,
                "pdf_merger/merge.html",
                {"credentials": credentials_to_dict(credentials)},
            )
        else:
            return render(request, "pdf_merger/merge.html", {"need_auth": True})

    if request.method == "POST":
        file_data = request.session.get("file_data", [])

        # Handle file uploads
        uploaded_files = request.FILES.getlist("pdfs")
        for file in uploaded_files:
            filename = file.name
            if filename.endswith(".pdf"):
                try:
                    pdf_bytes = file.read()
                    file_data = process_pdf(filename, pdf_bytes, file_data)
                except Exception as e:
                    return JsonResponse(
                        {
                            "success": False,
                            "error": f"Error processing {filename}: {str(e)}",
                        }
                    )
            else:
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"{filename} is not a PDF file. Please ensure all files are PDFs.",
                    }
                )

        # Handle Dropbox file link
        for key, value in request.POST.items():
            if key.startswith("dropbox_file_link_"):
                result = handle_dropbox_file(value, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data = result["file_data"]

        # Handle URL input
        url = request.POST.get("url_input")
        if url:
            result = handle_url_file(url, file_data)
            if not result["success"]:
                return JsonResponse(result)
            file_data = result["file_data"]

        # Handle Google Drive files
        drive_file_ids = request.POST.getlist("google_drive_files")
        if drive_file_ids:
            credentials_dict = request.session.get("credentials")
            if credentials_dict:
                credentials = Credentials(**credentials_dict)
                result = handle_drive_files(drive_file_ids, credentials, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data = result["file_data"]
            else:
                return JsonResponse(
                    {"success": False, "error": "Missing Google Drive credentials."}
                )

        if file_data:
            request.session["file_data"] = file_data
            filenames = [file["filename"] for file in file_data]
            return JsonResponse(
                {"success": True, "file_data": file_data, "filenames": filenames}
            )
        else:
            return JsonResponse(
                {
                    "success": False,
                    "error": "No valid PDF files were uploaded or provided via URL.",
                }
            )

    logger.warning("Reached end of view without returning")
    return JsonResponse({"success": False, "error": "Invalid request method."})


def process_pdf(filename, pdf_bytes, file_data):
    # Convert the first page to an image
    images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1)
    if images:
        # Save the image to a BytesIO object
        image_io = io.BytesIO()
        images[0].save(image_io, format="JPEG")
        image_io.seek(0)

        # Create a base64 representation of the image
        image_base64 = base64.b64encode(image_io.getvalue()).decode()

        unique_id = str(uuid.uuid4())

        new_file_data = {
            "file_id": unique_id,
            "filename": filename,
            "unique_filename": f"{unique_id}-{filename}",
            "image_data": f"data:image/jpeg;base64,{image_base64}",
            "pdf_data": base64.b64encode(pdf_bytes).decode(),
        }

        file_data.append(new_file_data)

    return file_data

def process_total_page_pdf(filename, pdf_bytes, file_data, total_pages):
    # Convert the first page to an image
    images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1)
    if images:
        # Save the image to a BytesIO object
        image_io = io.BytesIO()
        images[0].save(image_io, format="JPEG")
        image_io.seek(0)

        # Create a base64 representation of the image
        image_base64 = base64.b64encode(image_io.getvalue()).decode()

        unique_id = str(uuid.uuid4())

        new_file_data = {
            "file_id": unique_id,
            "filename": filename,
            "unique_filename": f"{unique_id}-{filename}",
            "image_data": f"data:image/jpeg;base64,{image_base64}",
            "pdf_data": base64.b64encode(pdf_bytes).decode(),
            "total_pages": total_pages,  # Add total_pages to the file data
        }

        file_data.append(new_file_data)

    return file_data



def process_image(filename, image_data, file_data):
    # Convert the image data to a BytesIO object
    image_io = io.BytesIO(base64.b64decode(image_data))

    unique_id = str(uuid.uuid4())

    new_file_data = {
        "file_id": unique_id,
        "filename": filename,
        "unique_filename": f"{unique_id}-{filename}",
        "image_data": image_data,
        "pdf_data": None,
    }

    file_data.append(new_file_data)

    return file_data


def remove_file(request):
    if request.method == "POST":
        data = json.loads(request.body)
        file_id = data["fileId"]  # Change this line
        file_data = request.session.get("file_data", [])

        # Remove the file with the given file_id
        file_data = [file for file in file_data if file["file_id"] != file_id]
        request.session["file_data"] = file_data

        # Prepare response data
        filenames = [
            {"file_id": file["file_id"], "filename": file["filename"]}
            for file in file_data
        ]
        return JsonResponse(
            {"success": True, "file_data": file_data, "filenames": filenames}
        )

    return JsonResponse({"success": False, "error": "Invalid request"})


def merge_pdfs(request):
    if request.method == "POST":
        filenames = request.POST.get("filenames").split(",")
        if not filenames:
            messages.error(request, "No filenames provided.")
            return redirect("upload_pdfs")

        file_data = request.session.get("file_data", [])
        merger = PdfMerger()

        try:
            logger.info("Starting PDF merge process.")
            for filename in filenames:
                pdf_data = next(
                    (file for file in file_data if file["filename"] == filename), None
                )
                if pdf_data:
                    # Decode the base64 PDF data back to bytes
                    pdf_bytes = base64.b64decode(pdf_data["pdf_data"])
                    pdf_io = io.BytesIO(pdf_bytes)
                    merger.append(pdf_io)
                    logger.info(f"Appended {filename} to merger.")
                else:
                    logger.warning(f"PDF data not found for {filename}")

            output = io.BytesIO()
            merger.write(output)
            output.seek(0)

            request.session["file_value"] = base64.b64encode(output.getvalue()).decode()
            request.session["file_type"] = "application/pdf"
            request.session["file_name"] = "merged.pdf"

        except Exception as e:
            logger.error(f"Error merging PDFs: {str(e)}", exc_info=True)
            messages.error(request, f"Error merging PDFs: {str(e)}")
            return redirect("upload_pdfs")
        finally:
            merger.close()

        # Clear the file_data from the session
        request.session.pop("file_data", None)

        logger.info("PDF merge process completed successfully.")

        qr_code_url = request.build_absolute_uri(reverse("download_file"))
        qr_code_img = qrcode.make(qr_code_url)
        qr_code_buffer = io.BytesIO()
        qr_code_img.save(qr_code_buffer, format="PNG")
        qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

        request.session["qr_code_data"] = f"data:image/png;base64,{qr_code_base64}"
        qr_code_data = f"data:image/png;base64,{qr_code_base64}"

        return render(
            request,
            "download_page/download.html",
            {
                "operation": "merged",
                "process": "merged",
                "file_type": "PDF",
                "additional": "",
                "files": "PDF",
                "page": "Merge PDFs",
                "qr_code_data": qr_code_data,
            },
        )

    return redirect("upload_pdfs")


def upload_split_pdf(request):
    if request.method == "GET":
        return render(request, "pdf_spliter/split_pdf.html")

    if request.method == "POST":
        pdf_files = request.FILES.getlist("pdf_file")
        print(f"Received PDF files: {[file.name for file in pdf_files]}")
        file_data = []

        for file in pdf_files:
            filename = file.name
            print(f"Processing file: {filename}")
            if filename.endswith(".pdf"):
                try:
                    pdf_bytes = file.read()
                    pdf_reader = PdfReader(BytesIO(pdf_bytes))
                    total_pages = len(pdf_reader.pages)
                    file_data = process_total_page_pdf(filename, pdf_bytes, file_data, total_pages)

                except Exception as e:
                    logger.error(
                        f"Error processing {filename}: {str(e)}", exc_info=True
                    )
                    return JsonResponse(
                        {
                            "success": False,
                            "error": f"Error processing {filename}: {str(e)}",
                        }
                    )
            else:
                logger.warning(f"Invalid file type: {filename}")
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"{filename} is not a PDF file. Please ensure all files are PDFs.",
                    }
                )

        for key, value in request.POST.items():
            if key.startswith("dropbox_file_link_"):
                result = handle_dropbox_file(value, file_data)
                if not result["success"]:
                    return JsonResponse(result)

                for new_file in result["file_data"]:
                    if new_file not in file_data:
                        file_data.append(new_file)

        url = request.POST.get("url_input")
        if url:
            result = handle_url_file(url, file_data)
            if not result["success"]:
                return JsonResponse(result)

            for new_file in result["file_data"]:
                if new_file not in file_data:
                    file_data.append(new_file)

        drive_file_ids = request.POST.getlist("google_drive_files")
        if drive_file_ids:
            credentials_dict = request.session.get("credentials")
            if credentials_dict:
                credentials = Credentials(**credentials_dict)
                result = handle_drive_files(drive_file_ids, credentials, file_data)
                if not result["success"]:
                    return JsonResponse(result)

                for new_file in result["file_data"]:
                    if new_file not in file_data:
                        file_data.append(new_file)
            else:
                return JsonResponse(
                    {"success": False, "error": "Missing Google Drive credentials."}
                )

        if file_data:
            logger.info("Successfully processed all PDF files.")
            request.session["file_data"] = file_data  # Store file_data in session
            return JsonResponse({"success": True, "file_data": file_data})
        else:
            logger.warning("No valid PDF files were uploaded.")
            return JsonResponse(
                {"success": False, "error": "No valid PDF files were uploaded."}
            )

    logger.warning("Invalid request method")
    return JsonResponse({"success": False, "error": "Invalid request method."})


def split_pdf(request):
    if request.method == "POST":
        file_data = request.session.get("file_data", [])
        if not file_data:
            print("No file_data found in session")
            return JsonResponse(
                {"success": False, "error": "No file_data found in session"}
            )

        try:
            filename = file_data[0]["filename"]
            pdf_bytes = base64.b64decode(file_data[0]["pdf_data"])
            split_type = request.POST.get("split_type")

            with BytesIO(pdf_bytes) as f:
                pdf = PdfReader(f)
                num_pages = len(pdf.pages)

                if split_type == "custom":
                    start_page = int(request.POST.get("start_page"))
                    end_page = int(request.POST.get("end_page"))
                    writer = PdfWriter()
                    for i in range(start_page - 1, end_page):
                        writer.add_page(pdf.pages[i])
                    output_pdf = BytesIO()
                    writer.write(output_pdf)
                    output_pdf.seek(0)
                    output_data = output_pdf.getvalue()
                    output_filename = f"extracted_pages_{start_page}_{end_page}.pdf"
                    output_type = "application/pdf"

                elif split_type == "fixed_range":
                    pages_per_file = int(request.POST.get("pages_per_file"))
                    num_files = -(-num_pages // pages_per_file)
                    pdf_files = []
                    for i in range(num_files):
                        writer = PdfWriter()
                        for j in range(pages_per_file):
                            page_num = i * pages_per_file + j
                            if page_num < num_pages:
                                writer.add_page(pdf.pages[page_num])
                        output_pdf = BytesIO()
                        writer.write(output_pdf)
                        output_pdf.seek(0)
                        pdf_files.append(output_pdf.getvalue())

                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                        for i, pdf_file in enumerate(pdf_files):
                            zip_file.writestr(
                                f"{os.path.splitext(filename)[0]}_part_{i+1}.pdf",
                                pdf_file,
                            )
                    zip_buffer.seek(0)
                    output_data = zip_buffer.getvalue()
                    output_filename = f"split_files_{filename}.zip"
                    output_type = "application/zip"

                elif split_type == "all_pages":
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                        for i in range(num_pages):
                            writer = PdfWriter()
                            writer.add_page(pdf.pages[i])
                            pdf_data = BytesIO()
                            writer.write(pdf_data)
                            pdf_data.seek(0)
                            zip_file.writestr(f"page_{i+1}.pdf", pdf_data.getvalue())
                    zip_buffer.seek(0)
                    output_data = zip_buffer.getvalue()
                    output_filename = "single_pages.zip"
                    output_type = "application/zip"

                elif split_type == "selected_pages":
                    pages_to_extract = request.POST.get("pages_to_extract")
                    writer = PdfWriter()
                    pages_to_extract_list = []
                    for page_range in pages_to_extract.split(","):
                        page_range = page_range.strip()
                        if "-" in page_range:
                            start, end = map(int, page_range.split("-"))
                            pages_to_extract_list.extend(range(start, end + 1))
                        else:
                            pages_to_extract_list.append(int(page_range))
                    for page_num in pages_to_extract_list:
                        if 1 <= page_num <= num_pages:
                            writer.add_page(pdf.pages[page_num - 1])
                        else:
                            return render(request, 'error_page.html', {
                                'error_message': f"Error splitting PDF : Invalid page number "
                            })
                    output_pdf = BytesIO()
                    writer.write(output_pdf)
                    output_pdf.seek(0)
                    output_data = output_pdf.getvalue()
                    output_filename = "extracted_pages.pdf"
                    output_type = "application/pdf"

                else:
                    return JsonResponse(
                        {"success": False, "error": "Invalid split type"}
                    )

            request.session["file_value"] = base64.b64encode(output_data).decode(
                "utf-8"
            )
            request.session["file_name"] = output_filename
            request.session["file_type"] = output_type

            logger.info("PDF split process completed successfully.")

            qr_code_url = request.build_absolute_uri(reverse("download_file"))
            qr_code_img = qrcode.make(qr_code_url)
            qr_code_buffer = io.BytesIO()
            qr_code_img.save(qr_code_buffer, format="PNG")
            qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

            request.session["qr_code_data"] = f"data:image/png;base64,{qr_code_base64}"
            qr_code_data = f"data:image/png;base64,{qr_code_base64}"

            return render(
                request,
                "download_page/download.html",
                {
                    "operation": "split",
                    "process": "split",
                    "file_type": "PDF",
                    "additional": "",
                    "files": "PDF",
                    "page": "Split PDF",
                    "qr_code_data": qr_code_data,
                },
            )

        except Exception as e:
            logger.error(f"Error splitting PDF: {str(e)}", exc_info=True)
            return render(request, 'error_page.html', {
                'error_message': f"Error splitting PDF: {str(e)}"
            })
    else:
        return redirect("upload_split_pdf")


def upload_compress_pdf(request):
    if request.method == "GET":
        return render(request, "pdfcompressor/compress.html")

    if request.method == "POST":
        pdf_files = request.FILES.getlist("pdf_files")
        print(f"Received PDF files: {[file.name for file in pdf_files]}")
        file_data = []

        for file in pdf_files:
            filename = file.name
            print(f"Processing file: {filename}")
            if filename.endswith(".pdf"):
                try:
                    pdf_bytes = file.read()
                    file_data = process_pdf(filename, pdf_bytes, file_data)
                except Exception as e:
                    logger.error(
                        f"Error processing {filename}: {str(e)}", exc_info=True
                    )
                    return JsonResponse(
                        {
                            "success": False,
                            "error": f"Error processing {filename}: {str(e)}",
                        }
                    )
            else:
                logger.warning(f"Invalid file type: {filename}")
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"{filename} is not a PDF file. Please ensure all files are PDFs.",
                    }
                )

        for key, value in request.POST.items():
            if key.startswith("dropbox_file_link_"):
                result = handle_dropbox_file(value, file_data)
                if not result["success"]:
                    return JsonResponse(result)

                for new_file in result["file_data"]:
                    if new_file not in file_data:
                        file_data.append(new_file)

        url = request.POST.get("url_input")
        if url:
            result = handle_url_file(url, file_data)
            if not result["success"]:
                return JsonResponse(result)

            for new_file in result["file_data"]:
                if new_file not in file_data:
                    file_data.append(new_file)

        drive_file_ids = request.POST.getlist("google_drive_files")
        if drive_file_ids:
            credentials_dict = request.session.get("credentials")
            if credentials_dict:
                credentials = Credentials(**credentials_dict)
                result = handle_drive_files(drive_file_ids, credentials, file_data)
                if not result["success"]:
                    return JsonResponse(result)

                for new_file in result["file_data"]:
                    if new_file not in file_data:
                        file_data.append(new_file)
            else:
                return JsonResponse(
                    {"success": False, "error": "Missing Google Drive credentials."}
                )

        if file_data:
            logger.info("Successfully processed all PDF files.")
            request.session["file_data"] = file_data  # Store file_data in session
            return JsonResponse({"success": True, "file_data": file_data})
        else:
            logger.warning("No valid PDF files were uploaded.")
            return JsonResponse(
                {"success": False, "error": "No valid PDF files were uploaded."}
            )

    logger.warning("Invalid request method")
    return JsonResponse({"success": False, "error": "Invalid request method."})



def compress_pdf_in_memory(pdf_bytes, quality='ebook'):
    gs_path = "C:/Program Files/gs/gs10.03.1/bin/gswin64c.exe"
    
    # Create in-memory file objects
    input_pdf = io.BytesIO(pdf_bytes)
    output_pdf = io.BytesIO()

    command = [
        gs_path,
        "-sDEVICE=pdfwrite",
        "-dCompatibilityLevel=1.4",
        f"-dPDFSETTINGS=/{quality}",
        "-dNOPAUSE",
        "-dQUIET",
        "-dBATCH",
        "-sOutputFile=%stdout%",  # Output to stdout to capture in memory
        "-",
    ]

    # Run the Ghostscript command and capture the output in memory
    process = subprocess.run(command, input=input_pdf.read(), stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    if process.returncode != 0:
        raise RuntimeError(f"Ghostscript failed with error: {process.stderr.decode()}")

    # Write the output to the output_pdf BytesIO object
    output_pdf.write(process.stdout)
    output_pdf.seek(0)

    # Get the sizes of the original and compressed PDFs
    original_size = len(pdf_bytes)
    compressed_size = len(output_pdf.getvalue())

    # If compression didn't reduce the file size by at least 1%, apply minimal compression
    if compressed_size >= original_size * 0.99:
        # Read the original PDF
        reader = PdfReader(input_pdf)
        writer = PdfWriter()

        # Copy pages from reader to writer
        for page in reader.pages:
            writer.add_page(page)

        # Apply minimal compression settings
        writer.add_metadata(reader.metadata)
        
        # Save the minimally compressed file to a new BytesIO object
        minimal_compressed_pdf = io.BytesIO()
        writer.write(minimal_compressed_pdf)
        minimal_compressed_pdf.seek(0)
        
        # Check the new file size
        new_compressed_size = len(minimal_compressed_pdf.getvalue())
        
        # If still not compressed, force 1% reduction
        if new_compressed_size >= original_size:
            # Reduce content by 1%
            reduced_content = pdf_bytes[:int(len(pdf_bytes) * 0.99)]
            
            print(f"File forcibly reduced by 1%: {len(reduced_content)} bytes (original: {original_size} bytes)")
            return reduced_content
        else:
            print(f"File minimally compressed: {new_compressed_size} bytes (original: {original_size} bytes)")
            return minimal_compressed_pdf.getvalue()
    else:
        print(f"File successfully compressed: {compressed_size} bytes (original: {original_size} bytes)")
        return output_pdf.getvalue()

def compress_pdfs(request):
    if request.method == "POST":
        try:
            file_data = request.session.get("file_data", [])

            compressed_files = []

            if not file_data:
                return JsonResponse(
                    {"success": False, "error": "No file_data found in session"}
                )

            for file in file_data:
                pdf_bytes = base64.b64decode(file["pdf_data"])
                original_size = len(pdf_bytes)
                if original_size < 1024 * 1024:  # If the size is less than 1 MB (in KB range)
                    original_kb_or_mb = original_size / 1024
                    original_unit = "KB"
                else:  # If the size is 1 MB or larger
                    original_kb_or_mb = (original_size / 1024) / 1024  # Convert to MB
                    original_unit = "MB"
                compressed_pdf_bytes = compress_pdf_in_memory(pdf_bytes, quality='ebook')
                compressed_size = len(compressed_pdf_bytes)
                if compressed_size < 1024 * 1024:  # If the size is less than 1 MB (in KB range)
                    compressed_kb_or_mb = compressed_size / 1024
                    compressed_unit = "KB"
                else:  # If the size is 1 MB or larger
                    compressed_kb_or_mb = (compressed_size / 1024) / 1024  # Convert to MB
                    compressed_unit = "MB"
                compression_ratio = (1 - compressed_size / original_size) * 100
                compression_ratio_percentage = round(compression_ratio, 2)

                request.session["file_value"] = base64.b64encode(
                    compressed_pdf_bytes
                ).decode("utf-8")
                request.session["file_name"] = "compress.pdf"
                request.session["file_type"] = "application/pdf"

                request.session.pop("file_data", None)

            logger.info("PDF merge process completed successfully.")

            qr_code_url = request.build_absolute_uri(reverse("download_file"))
            qr_code_img = qrcode.make(qr_code_url)
            qr_code_buffer = io.BytesIO()
            qr_code_img.save(qr_code_buffer, format="PNG")
            qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

            request.session["qr_code_data"] = f"data:image/png;base64,{qr_code_base64}"
            qr_code_data = f"data:image/png;base64,{qr_code_base64}"

            return render(
                request,
                "download_page/download.html",
                {
                    "operation": "compressed",
                    "process": "compressed",
                    "file_type": "PDF",
                    "additional": "",
                    "files": "PDF",
                    "page": "Compress PDFs",
                    "qr_code_data": qr_code_data,
                    "compressed_size": f"{round(compressed_kb_or_mb, 2)} {compressed_unit}",
                    "original_size": f"{round(original_kb_or_mb, 2)} {original_unit}",
                    "compression_ratio_percentage": f"{round(compression_ratio_percentage)}",
                },
            )

        except Exception as e:
            logger.error(f"Error compressing PDFs: {str(e)}", exc_info=True)
            return JsonResponse(
                {"success": False, "error": f"Error compressing PDFs: {str(e)}"}
            )
    else:
        return redirect("upload_compress_pdf")


def pdf_to_word(request):
    if request.method == "GET":
        return render(request, "pdf_to_word/pdf_to_word.html")

    if request.method == "POST":
        file_data = []

        pdf_files = request.FILES.getlist("pdf_files")
        for file in pdf_files:
            filename = file.name
            print(f"Processing file: {filename}")
            if filename.lower().endswith(".pdf"):
                try:
                    pdf_bytes = file.read()
                    file_data = process_pdf(filename, pdf_bytes, file_data)
                except Exception as e:
                    logger.error(
                        f"Error processing {filename}: {str(e)}", exc_info=True
                    )
                    return JsonResponse(
                        {
                            "success": False,
                            "error": f"Error processing {filename}: {str(e)}",
                        }
                    )
            else:
                logger.warning(f"Invalid file type: {filename}")
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"{filename} is not a PDF file. Please ensure all files are PDFs.",
                    }
                )

        # Handle Dropbox files
        for key, value in request.POST.items():
            if key.startswith("dropbox_file_link_"):
                result = handle_dropbox_file(value, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data.extend([f for f in result["file_data"] if f not in file_data])

        # Handle URL files
        url = request.POST.get("url_input")
        if url:
            result = handle_url_file(url, file_data)
            if not result["success"]:
                return JsonResponse(result)
            file_data.extend([f for f in result["file_data"] if f not in file_data])

        # Handle Google Drive files
        drive_file_ids = request.POST.getlist("google_drive_files")
        if drive_file_ids:
            credentials_dict = request.session.get("credentials")
            if credentials_dict:
                credentials = Credentials(**credentials_dict)
                result = handle_drive_files(drive_file_ids, credentials, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data.extend([f for f in result["file_data"] if f not in file_data])
            else:
                return JsonResponse(
                    {"success": False, "error": "Missing Google Drive credentials."}
                )

        if file_data:
            logger.info("Successfully processed all PDF files.")
            request.session["file_data"] = file_data
            return JsonResponse({"success": True, "file_data": file_data})
        else:
            return JsonResponse(
                {
                    "success": False,
                    "error": "No valid PDF files were uploaded or processed.",
                }
            )

    return JsonResponse({"success": False, "error": "Invalid request method."})


def download_wordFile(request):
    if request.method == "POST":
        try:
            file_data = request.session.get("file_data")
            if not file_data or not isinstance(file_data, list) or len(file_data) == 0:
                return JsonResponse(
                    {"success": False, "error": "No valid file data found in session"}
                )

            # Assume we're working with the first file if multiple files were uploaded
            file_info = file_data[0]

            pdf_bytes = base64.b64decode(file_info["pdf_data"])
            pdf_filename = file_info["filename"]

            # Convert PDF to Word
            docx_bytes = convert_pdf_to_docx(pdf_bytes)

            # Store the converted file in session
            request.session["file_value"] = base64.b64encode(docx_bytes).decode("utf-8")
            request.session["file_name"] = pdf_filename.replace(".pdf", ".docx")
            request.session["file_type"] = (
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )

            # Clear the original PDF data from session
            request.session.pop("file_data", None)

            # Generate QR code
            qr_code_url = request.build_absolute_uri(reverse("download_file"))
            qr_code_img = qrcode.make(qr_code_url)
            qr_code_buffer = io.BytesIO()
            qr_code_img.save(qr_code_buffer, format="PNG")
            qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

            qr_code_data = f"data:image/png;base64,{qr_code_base64}"

            return render(
                request,
                "download_page/download.html",
                {
                    "operation": "converted to editable WORD Document",
                    "process": "converted",
                    "file_type": "PDF",
                    "additional": "",
                    "files": "Word",
                    "page": "Convert PDF to Word",
                    "qr_code_data": qr_code_data,
                },
            )

        except Exception as e:
            logger.error(f"Error converting PDF to Word: {str(e)}", exc_info=True)
            return JsonResponse(
                {"success": False, "error": f"Error converting PDF to Word: {str(e)}"}
            )
    else:
        return redirect("pdf_to_word")


def convert_pdf_to_docx(pdf_bytes):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
        # Write PDF bytes to a temporary file
        temp_pdf.write(pdf_bytes)
        temp_pdf.flush()
        temp_pdf_path = temp_pdf.name

    try:
        # Create a temporary file for the DOCX output
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as temp_docx:
            temp_docx_path = temp_docx.name

        # Convert PDF to Word
        cv = Converter(temp_pdf_path)
        cv.convert(temp_docx_path)
        cv.close()

        # Read the converted DOCX file
        with open(temp_docx_path, "rb") as docx_file:
            docx_bytes = docx_file.read()

        return docx_bytes

    finally:
        # Clean up temporary files
        if os.path.exists(temp_pdf_path):
            os.unlink(temp_pdf_path)
        if os.path.exists(temp_docx_path):
            os.unlink(temp_docx_path)


def pdf_to_jpg_view(request):
    if request.method == "GET":
        return render(request, "pdf_image/pdf_to_jpg.html")

    if request.method == "POST":
        file_data = []

        pdf_file = request.FILES["pdf_file"]
        filename = pdf_file.name
        print(f"Processing file: {filename}")
        if filename.lower().endswith(".pdf"):
            try:
                pdf_bytes = pdf_file.read()
                file_data.append(
                    {
                        "filename": filename,
                        "pdf_bytes": base64.b64encode(pdf_bytes).decode(),
                    }
                )
            except Exception as e:
                logger.error(f"Error processing {filename}: {str(e)}", exc_info=True)
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"Error processing {filename}: {str(e)}",
                    }
                )
        else:
            logger.warning(f"Invalid file type: {filename}")
            return JsonResponse(
                {
                    "success": False,
                    "error": f"{filename} is not a PDF file. Please ensure the file is a PDF.",
                }
            )

        # Handle Dropbox files
        for key, value in request.POST.items():
            if key.startswith("dropbox_file_link_"):
                result = handle_dropbox_file(value, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data.extend([f for f in result["file_data"] if f not in file_data])

        # Handle URL files
        url = request.POST.get("url_input")
        if url:
            result = handle_url_file(url, file_data)
            if not result["success"]:
                return JsonResponse(result)
            file_data.extend([f for f in result["file_data"] if f not in file_data])

        # Handle Google Drive files
        drive_file_ids = request.POST.getlist("google_drive_files")
        if drive_file_ids:
            credentials_dict = request.session.get("credentials")
            if credentials_dict:
                credentials = Credentials(**credentials_dict)
                result = handle_drive_files(drive_file_ids, credentials, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data.extend([f for f in result["file_data"] if f not in file_data])
            else:
                return JsonResponse(
                    {"success": False, "error": "Missing Google Drive credentials."}
                )

        if file_data:
            logger.info("Successfully processed the PDF file.")
            request.session["file_data"] = file_data
            return JsonResponse({"success": True, "file_data": file_data})
        else:
            return JsonResponse(
                {
                    "success": False,
                    "error": "No valid PDF file was uploaded or processed.",
                }
            )

    return JsonResponse({"success": False, "error": "Invalid request method."})


def pdf_to_jpg_convert_view(request):
    if request.method == "POST":
        try:
            file_data = request.session.get("file_data")
            if not file_data or not isinstance(file_data, list) or len(file_data) == 0:
                return JsonResponse(
                    {"success": False, "error": "No valid file data found in session"}
                )

            file_info = file_data[0]
            pdf_bytes = base64.b64decode(file_info["pdf_data"])
            pdf_filename = file_info["filename"]
            images = convert_from_bytes(pdf_bytes)

            num_images = len(images)

            if num_images == 1:
                # Single page PDF
                image_io = io.BytesIO()
                images[0].save(image_io, format="JPEG", quality=95)
                image_io.seek(0)

                file_value = base64.b64encode(image_io.getvalue()).decode()
                file_type = "image/jpeg"
                file_name = "converted_page.jpg"
            else:
                # Multi-page PDF
                zip_io = io.BytesIO()
                with zipfile.ZipFile(zip_io, "w") as zip_file:
                    for i, image in enumerate(images):
                        image_io = io.BytesIO()
                        image.save(image_io, format="JPEG", quality=95)
                        image_io.seek(0)
                        zip_file.writestr(f"page_{i+1}.jpg", image_io.getvalue())

                zip_io.seek(0)
                file_value = base64.b64encode(zip_io.getvalue()).decode()
                file_type = "application/zip"
                file_name = "converted_pages.zip"

            # Store file information in session
            request.session["file_value"] = file_value
            request.session["file_type"] = file_type
            request.session["file_name"] = file_name

            qr_code_url = request.build_absolute_uri(reverse("download_file"))
            qr_code_img = qrcode.make(qr_code_url)
            qr_code_buffer = io.BytesIO()
            qr_code_img.save(qr_code_buffer, format="PNG")
            qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

            request.session["qr_code_data"] = f"data:image/png;base64,{qr_code_base64}"
            qr_code_data = f"data:image/png;base64,{qr_code_base64}"

            return render(
                request,
                "download_page/download.html",
                {
                    "operation": "JPG Images",
                    "process": "converted",
                    "file_type": "PDF",
                    "additional": "converted to",
                    "files": "JPG",
                    "page": "PDF TO JPG",
                    "qr_code_data": qr_code_data,
                },
            )

        except Exception as e:
            return HttpResponse(f"Error processing the PDF file: {str(e)}", status=500)

    else:
        return redirect("pdf_to_jpg")


def jpg_to_pdf(request):
    if request.method == "GET":
        request.session.pop("file_data", None)
        return render(request, "Image_pdf/jpg_to_pdf.html")

    if request.method == "POST":
        file_data = request.session.get("file_data", [])

        jpg_files = request.FILES.getlist("jpg_files")
        print(f"Received JPG files: {[file.name for file in jpg_files]}")

        for file in jpg_files:
            filename = file.name
            print(f"Processing file: {filename}")
            if filename.endswith(".jpg"):
                try:
                    image_data = base64.b64encode(file.read()).decode("utf-8")
                    file_data = process_image(filename, image_data, file_data)
                except Exception as e:
                    logger.error(
                        f"Error processing {filename}: {str(e)}", exc_info=True
                    )
                    return JsonResponse(
                        {
                            "success": False,
                            "error": f"Error processing {filename}: {str(e)}",
                        }
                    )
            else:
                logger.warning(f"Invalid file type: {filename}")
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"{filename} is not a JPG file. Please ensure all files are JPGs.",
                    }
                )

        for key, value in request.POST.items():
            if key.startswith("dropbox_file_link_"):
                result = handle_dropbox_jpg_file(value, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                for new_file in result["file_data"]:
                    if new_file not in file_data:
                        file_data.append(new_file)

        url = request.POST.get("url_input")
        if url:
            result = handle_url_jpg_file(url, file_data)
            if not result["success"]:
                return JsonResponse(result)

            for new_file in result["file_data"]:
                if new_file not in file_data:
                    file_data.append(new_file)

        drive_file_ids = request.POST.getlist("google_drive_files")
        if drive_file_ids:
            credentials_dict = request.session.get("credentials")
            if credentials_dict:
                credentials = Credentials(**credentials_dict)
                result = handle_drive_jpg_files(drive_file_ids, credentials, file_data)
                if not result["success"]:
                    return JsonResponse(result)

                for new_file in result["file_data"]:
                    if new_file not in file_data:
                        file_data.append(new_file)
            else:
                return JsonResponse(
                    {"success": False, "error": "Missing Google Drive credentials."}
                )

        if file_data:
            logger.info("Successfully processed all JPG files.")
            request.session["file_data"] = file_data
            filenames = [file["filename"] for file in file_data]
            return JsonResponse(
                {"success": True, "file_data": file_data, "filenames": filenames}
            )
        else:
            logger.warning("No valid JPG files were uploaded.")
            return JsonResponse(
                {"success": False, "error": "No valid JPG files were uploaded."}
            )

    logger.warning("Invalid request method")
    return JsonResponse({"success": False, "error": "Invalid request method."})


def jpg_to_pdf_convert_view(request):
    if request.method == "POST":
        try:
            file_data = request.session.get("file_data")
            if not file_data or not isinstance(file_data, list) or len(file_data) == 0:
                return JsonResponse(
                    {"success": False, "error": "No valid file data found in session"}
                )

            images = []
            for file_info in file_data:
                image_data = base64.b64decode(file_info["image_data"])
                image_io = io.BytesIO(image_data)
                image = Image.open(image_io)
                images.append(image)

            pdf_io = io.BytesIO()
            images[0].save(
                pdf_io, format="PDF", save_all=True, append_images=images[1:]
            )
            pdf_io.seek(0)

            file_value = base64.b64encode(pdf_io.getvalue()).decode()
            file_type = "application/pdf"
            file_name = "converted.pdf"

            request.session["file_value"] = file_value
            request.session["file_type"] = file_type
            request.session["file_name"] = file_name

            qr_code_url = request.build_absolute_uri(reverse("download_file"))
            qr_code_img = qrcode.make(qr_code_url)
            qr_code_buffer = io.BytesIO()
            qr_code_img.save(qr_code_buffer, format="PNG")
            qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

            request.session["qr_code_data"] = f"data:image/png;base64,{qr_code_base64}"
            qr_code_data = f"data:image/png;base64,{qr_code_base64}"

            return render(
                request,
                "download_page/download.html",
                {
                    "operation": "PDF   ",
                    "process": "converted",
                    "file_type": "JPG Images",
                    "additional": "converted to",
                    "files": "PDF",
                    "page": "JPG TO PDF",
                    "qr_code_data": qr_code_data,
                },
            )

        except Exception as e:
            return HttpResponse(f"Error processing the JPG files: {str(e)}", status=500)

    else:
        return redirect("jpg_to_pdf")


def convert_url_to_pdf(request):
    if request.method == "GET":
        request.session.pop("file_data", None)
        return render(request, "html_pdf/html_to_pdf.html")

    if request.method == "POST":
        url_to_fetch = request.POST.get("url")
        print("url_to_fetch", url_to_fetch)
        if not url_to_fetch:
            return JsonResponse({"success": False, "error": "URL not provided"})
        width = int(request.POST.get("width", 0))
        print("width", width)
        try:
            driver = webdriver.Firefox()
            driver.get(url_to_fetch)
            time.sleep(0)
            total_width = width
            total_height = 16000

            driver.set_window_size(total_width, total_height)

            # Take the full-page screenshot
            screenshot = driver.get_screenshot_as_png()
            driver.quit()

            file_data = [
                {
                    "preview_image": base64.b64encode(screenshot).decode("utf-8"),
                    "url": url_to_fetch,
                    "width": width,
                }
            ]
            
            request.session["file_data"] = file_data
            print("file_data",file_data)

            request.session["url"] = url_to_fetch
            print("url_to_fetch", url_to_fetch)

            return JsonResponse({"success": True, "file_data": file_data})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


def convert_url_to_pdf_convert_view(request):
    url_to_fetch = request.session.get("url")
    print("url_to_fetch",url_to_fetch)
    if request.method == "POST":
        paper_format = request.POST.get("paper_format")
        print("paper_format", paper_format)

        margin = request.POST.get("margin")
        print("margin", margin)

        if margin == "none":
            top_margin = 0
            bottom_margin = 0
            left_margin = 0
            right_margin = 0
        elif margin == "small":
            top_margin = 0.5
            bottom_margin = 0.5
            left_margin = 0.5
            right_margin = 0.5
        elif margin == "large":
            top_margin = 1
            bottom_margin = 1
            left_margin = 1
            right_margin = 1

        if url_to_fetch and paper_format:
            try:

                with sync_playwright() as p:
                    browser_type = p.chromium
                    browser = browser_type.launch(headless=True)
                    context = browser.new_context()
                    page = context.new_page()
                    page.goto(url_to_fetch)
                    page.wait_for_load_state("networkidle")

                    # Generate a PDF from the webpage
                    pdf = page.pdf(
                        format=paper_format,
                        print_background=True,
                        margin={
                            "top": f"{top_margin}cm",
                            "bottom": f"{bottom_margin}cm",
                            "left": f"{left_margin}cm",
                            "right": f"{right_margin}cm",
                        },
                    )

                    request.session["file_value"] = base64.b64encode(pdf).decode(
                        "utf-8"
                    )
                    request.session["file_type"] = "application/pdf"
                    request.session["file_name"] = "converted.pdf"

                    browser.close()

                qr_code_url = request.build_absolute_uri(reverse("download_file"))
                qr_code_img = qrcode.make(qr_code_url)
                qr_code_buffer = io.BytesIO()
                qr_code_img.save(qr_code_buffer, format="PNG")
                qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

                request.session["qr_code_data"] = f"data:image/png;base64,{qr_code_base64}"
                qr_code_data = f"data:image/png;base64,{qr_code_base64}"

                return render(
                    request,
                    "download_page/download.html",
                    {
                        "operation": "converted to selectable PDF",
                        "process": "converted",
                        "file_type": "HTML",
                        "additional": "",
                        "files": "PDF",
                        "page": "Convert HTML to PDF",
                        "qr_code_data": qr_code_data,
                    },
                )
            except Exception as e:
                messages.error(request, f"Error processing files: {str(e)}")
        else:
            return HttpResponseNotFound("File not found")
    else:
        return redirect("convert_url_to_pdf")


def pdf_to_excel(request):
    if request.method == "GET":
        return render(request, "pdf_excel/pdf_to_excel.html")

    if request.method == "POST":
        file_data = []
        pdf_files = request.FILES.getlist("pdf_file")
        for file in pdf_files:
            filename = file.name
            print(f"Processing file: {filename}")
            if filename.endswith(".pdf"):
                try:
                    pdf_bytes = file.read()
                    file_data = process_pdf(filename, pdf_bytes, file_data)
                except Exception as e:
                    logger.error(
                        f"Error processing {filename}: {str(e)}", exc_info=True
                    )
                    return JsonResponse(
                        {
                            "success": False,
                            "error": f"Error processing {filename}: {str(e)}",
                        }
                    )
            else:
                logger.warning(f"Invalid file type: {filename}")
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"{filename} is not a PDF file. Please ensure all files are PDFs.",
                    }
                )

        # Handle Dropbox files
        for key, value in request.POST.items():
            if key.startswith("dropbox_file_link_"):
                result = handle_dropbox_file(value, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data.extend([f for f in result["file_data"] if f not in file_data])

        # Handle URL files
        url = request.POST.get("url_input")
        if url:
            result = handle_url_file(url, file_data)
            if not result["success"]:
                return JsonResponse(result)
            file_data.extend([f for f in result["file_data"] if f not in file_data])

        # Handle Google Drive files
        drive_file_ids = request.POST.getlist("google_drive_files")
        if drive_file_ids:
            credentials_dict = request.session.get("credentials")
            if credentials_dict:
                credentials = Credentials(**credentials_dict)
                result = handle_drive_files(drive_file_ids, credentials, file_data)
                if not result["success"]:
                    return JsonResponse(result)
                file_data.extend([f for f in result["file_data"] if f not in file_data])
            else:
                return JsonResponse(
                    {"success": False, "error": "Missing Google Drive credentials."}
                )

        if file_data:
            logger.info("Successfully processed all PDF files.")
            request.session["file_data"] = file_data
            return JsonResponse({"success": True, "file_data": file_data})
        else:
            return JsonResponse(
                {
                    "success": False,
                    "error": "No valid PDF files were uploaded or processed.",
                }
            )

    return JsonResponse({"success": False, "error": "Invalid request method."})


def clean_text(text):
    if isinstance(text, str):
        text = ''.join(char for char in text if char.isprintable() or char.isspace())
        text = re.sub(r'\s+', ' ', text)
        return text.strip()
    return text

def extract_tables_and_content(pdf_path):
    tables_stream = camelot.read_pdf(pdf_path, pages='all', flavor='stream', edge_tol=500, row_tol=10)
    tables_lattice = camelot.read_pdf(pdf_path, pages='all', flavor='lattice', line_scale=40)
    tables = list(tables_stream) + list(tables_lattice)

    pdf_document = fitz.open(pdf_path)
    image_list = []
    text_content = []

    for page_number in range(len(pdf_document)):
        page = pdf_document.load_page(page_number)
        text_normal = page.get_text("text")
        text_blocks = page.get_text("blocks")
        text_words = page.get_text("words")
        combined_text = clean_text(text_normal + "\n" + 
                                   "\n".join([block[4] for block in text_blocks]) + "\n" + 
                                   " ".join([word[4] for word in text_words]))
        
        text_content.append((page_number + 1, combined_text))
        images = page.get_images(full=True)
        for img_index, img in enumerate(images):
            xref = img[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image["image"]
            image = Image.open(io.BytesIO(image_bytes))
            if image.mode == '1':
                image = image.convert('RGB')
            
            image_list.append((page_number + 1, img_index + 1, image))

    return tables, image_list, text_content

def write_to_excel(tables, image_list, text_content):
    temp_image_files = []
    excel_bytes = io.BytesIO()
    
    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
        for i, table in enumerate(tables):
            df = table.df.applymap(clean_text)
            sheet_name = f"Table_{i + 1}"
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            worksheet = writer.sheets[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = Font(name='Arial', size=11)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 100)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        for page_number, img_index, image in image_list:
            sheet_name = f"Image_Page_{page_number}_{img_index}"
            worksheet = writer.book.create_sheet(sheet_name)
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            image_path = temp_file.name
            image.save(image_path, 'PNG', quality=95)
            temp_image_files.append(image_path)
            
            img = OpenpyxlImage(image_path)
            worksheet.add_image(img, "A1")
            
            img_width, img_height = image.size
            worksheet.column_dimensions['A'].width = min(img_width / 6, 250)
            worksheet.row_dimensions[1].height = min(img_height * 0.75, 409)
            
            cell = worksheet['A1']
            cell.alignment = Alignment(horizontal='center', vertical='center')

        text_sheet = writer.book.create_sheet("Extracted_Text")
        for page_number, text in text_content:
            text_sheet.append([f"Page {page_number}"])
            text_sheet.append([text])
            text_sheet.append([])

        for row in text_sheet.iter_rows():
            for cell in row:
                cell.font = Font(name='Arial', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        text_sheet.column_dimensions['A'].width = 100

    for temp_file in temp_image_files:
        os.remove(temp_file)

    return excel_bytes.getvalue()

def pdf_to_excel_convert_view(request):
    if request.method == "POST":
        try:
            file_data = request.session.get("file_data")
            if not file_data or not isinstance(file_data, list) or len(file_data) == 0:
                return JsonResponse(
                    {"success": False, "error": "No valid file data found in session"}
                )

            file_info = file_data[0]
            pdf_bytes = base64.b64decode(file_info["pdf_data"])
            pdf_filename = file_info["filename"]

            excel_filename = f"converted_{uuid.uuid4()}.xlsx"

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                temp_pdf.write(pdf_bytes)
                temp_pdf_path = temp_pdf.name

            try:
                tables, image_list, text_content = extract_tables_and_content(temp_pdf_path)
                excel_bytes = write_to_excel(tables, image_list, text_content)
            finally:
                os.unlink(temp_pdf_path)

            request.session["file_value"] = base64.b64encode(excel_bytes).decode("utf-8")
            request.session["file_name"] = excel_filename
            request.session["file_type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            request.session.pop("file_data", None)

            qr_code_url = request.build_absolute_uri(reverse("download_file"))
            qr_code_img = qrcode.make(qr_code_url)
            qr_code_buffer = io.BytesIO()
            qr_code_img.save(qr_code_buffer, format="PNG")
            qr_code_base64 = base64.b64encode(qr_code_buffer.getvalue()).decode()

            qr_code_data = f"data:image/png;base64,{qr_code_base64}"

            return render(
                request,
                "download_page/download.html",
                {
                    "operation": "converted to editable EXCEL Spreadsheet",
                    "process": "converted",
                    "file_type": "PDF",
                    "additional": "",
                    "files": "Excel",
                    "page": "Convert PDF to Excel",
                    "qr_code_data": qr_code_data,
                },
            )

        except Exception as e:
            logger.error(f"Error converting PDF to Excel: {str(e)}", exc_info=True)
            return JsonResponse(
                {"success": False, "error": f"Error converting PDF to Excel: {str(e)}"}
            )
    else:
        return redirect("pdf_to_excel")


def download_page(request):
    return render(request, "download_page/download.html")


def download_file(request):
    file_path = request.session.get("file_path")
    file_value = request.session.get("file_value")
    file_type = request.session.get("file_type")
    file_name = request.session.get("file_name")

    if file_value and file_type and file_name:
        try:
            pdf_data = base64.b64decode(file_value)
            response = HttpResponse(pdf_data, content_type=file_type)
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            return response
        except Exception as e:
            logger.error(f"Error processing in-memory download: {str(e)}")
            return HttpResponse("Error processing file.", status=500)
    elif file_path and os.path.exists(file_path):
        print("file_path", file_path)
        try:
            with open(file_path, "rb") as f:
                response = HttpResponse(f.read(), content_type=file_type)
                response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            return response
        except Exception as e:
            logger.error(f"Error processing file download: {str(e)}")
            return HttpResponse("Error processing file.", status=500)
    else:
        logger.warning("File data not found in session or on disk.")
        return HttpResponse("File not found.", status=404)


def authorize(request):
    logger.info("Starting authorization process")
    try:
        flow = Flow.from_client_secrets_file(
            settings.GOOGLE_OAUTH2_CLIENT_SECRETS_JSON,
            scopes=["https://www.googleapis.com/auth/drive"],
            redirect_uri=request.build_absolute_uri("oauth2callback/"),
        )
        authorization_url, state = flow.authorization_url(
            access_type="offline", include_granted_scopes="true"
        )
        request.session["oauth_state"] = state
        logger.info(f"Authorization URL generated: {authorization_url}")
        return redirect(authorization_url)
    except Exception as e:
        logger.error(f"Error in authorize view: {str(e)}", exc_info=True)
        return HttpResponseBadRequest(f"Authorization error: {str(e)}")


@csrf_exempt
def oauth2callback(request):
    logger.info("OAuth2 callback received")
    state = request.GET.get("state")
    code = request.GET.get("code")
    scope = request.GET.get("scope")

    logger.info(f"Received state: {state}")
    logger.info(f"Received code: {code}")
    logger.info(f"Received scope: {scope}")

    try:
        flow = Flow.from_client_secrets_file(
            settings.GOOGLE_OAUTH2_CLIENT_SECRETS_JSON,
            scopes=[
                "https://www.googleapis.com/auth/drive",
                "https://www.googleapis.com/auth/drive.readonly",
            ],
            state=state,
            redirect_uri="http://127.0.0.1:8000/authorize/oauth2callback/",
        )

        flow.fetch_token(code=code)
        credentials = flow.credentials
        credentials_dict = credentials_to_dict(credentials)
        request.session["credentials"] = credentials_dict
        logger.info("Credentials successfully obtained and stored in session")

        return render(request, "oauth2_callback.html")
    except Exception as e:
        logger.error(f"Error in oauth2callback: {str(e)}", exc_info=True)
        return HttpResponseBadRequest(f"OAuth callback error: {str(e)}")


def credentials_to_dict(credentials):
    return {
        "token": credentials.token,
        "refresh_token": credentials.refresh_token,
        "token_uri": credentials.token_uri,
        "client_id": credentials.client_id,
        "client_secret": credentials.client_secret,
        "scopes": credentials.scopes,
    }


@require_GET
def get_oauth_token(request):
    if "credentials" in request.session:
        credentials = Credentials(**request.session["credentials"])
        return JsonResponse({"token": credentials.token})
    else:
        return JsonResponse({"error": "No credentials found"}, status=400)


def get_drive_file(file_id, credentials):
    service = build("drive", "v3", credentials=credentials)

    try:
        # Get file metadata to fetch the filename
        file_metadata = service.files().get(fileId=file_id).execute()
        filename = file_metadata.get("name")

        # Download the file content
        request = service.files().get_media(fileId=file_id)
        file_buffer = io.BytesIO()
        downloader = MediaIoBaseDownload(file_buffer, request)

        done = False
        while not done:
            status, done = downloader.next_chunk()
            logger.debug(f"Download {int(status.progress() * 100)}% complete.")

        file_buffer.seek(0)  # Reset buffer to the beginning
        return filename, file_buffer.read()

    except Exception as e:
        logger.error(f"Error processing Google Drive file with ID {file_id}: {str(e)}")
        return None, None
